"""
処方箋Excelを public/data/prescriptions.json に変換するスクリプト。
使い方:
  変換: py excel_to_prescriptions_json.py
  タブ一覧のみ表示: py excel_to_prescriptions_json.py --list
"""

import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl がインストールされていません。次のコマンドを実行してください:")
    print("  pip install openpyxl")
    exit(1)

EXCEL_NAME = "ライフオラクル_処方箋_全職種マスター.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "public", "data", "prescriptions.json")

# Excel の列名候補（シートによって違う場合に対応）
COL_タイプID = ("タイプID", "typeId")
COL_年代 = ("年代", "age")
COL_本文 = ("本文", "text")

# タブ名 → アプリの職種ID（タブ名にスラッシュが使えないため）
SHEET_TO_JOB = {
    "主婦主夫": "主婦/主夫",
}


def list_sheets():
    """Excelのタブ（シート）名を一覧表示する"""
    excel_path = os.path.join(os.path.dirname(__file__), EXCEL_NAME)
    if not os.path.exists(excel_path):
        print(f"エラー: {EXCEL_NAME} が見つかりません。")
        print(f"次の場所にファイルを置いてください: {excel_path}")
        exit(1)
    wb = load_workbook(excel_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    print(f"【{EXCEL_NAME} のタブ一覧】全 {len(names)} 枚")
    print("-" * 40)
    for i, name in enumerate(names, 1):
        print(f"  {i:2}. {name}")
    print("-" * 40)
    print("※ タブ名がそのまま「職種」としてJSONのキーに使われます。")
    print("  アプリの職種と揃えると処方箋が正しく表示されます。")


def find_col_index(header_row, *candidates):
    """ヘッダー行から列名に合う列インデックス（1-based）を返す"""
    for cell in header_row:
        if cell.value is None:
            continue
        v = str(cell.value).strip()
        for c in candidates:
            if c == v:
                return cell.column
    return None


def main():
    excel_path = os.path.join(os.path.dirname(__file__), EXCEL_NAME)
    if not os.path.exists(excel_path):
        print(f"エラー: {EXCEL_NAME} が見つかりません。")
        print(f"次の場所にファイルを置いてください: {excel_path}")
        exit(1)

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    result = {}

    # タブ（シート）ごとに処理。タブ名 = 職種（アプリと揃えるためマッピングあり）
    for sheet_name in wb.sheetnames:
        職種 = SHEET_TO_JOB.get(sheet_name.strip(), sheet_name.strip())
        ws = wb[sheet_name]

        header = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not header:
            continue
        header_row = list(header) if isinstance(header, (list, tuple)) else [header]

        col_タイプID = find_col_index(header_row, *COL_タイプID)
        col_年代 = find_col_index(header_row, *COL_年代)
        col_本文 = find_col_index(header_row, *COL_本文)

        if not all([col_タイプID, col_年代, col_本文]):
            print(f"  スキップ: シート「{sheet_name}」にタイプID・年代・本文の列がありません")
            continue

        for row in ws.iter_rows(min_row=2):
            try:
                タイプID = row[col_タイプID - 1].value
                年代 = row[col_年代 - 1].value
                本文 = row[col_本文 - 1].value
            except IndexError:
                continue
            if not all([タイプID, 年代, 本文]):
                continue
            タイプID = str(タイプID).strip()
            年代 = str(年代).strip()
            本文 = str(本文).strip() if 本文 else ""
            key = f"{職種}_{タイプID}_{年代}"
            result[key] = {"text": 本文}

    wb.close()

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"完了: {len(result)} 件を {OUTPUT_PATH} に出力しました。")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] in ("--list", "-l", "/list"):
        list_sheets()
    else:
        main()
